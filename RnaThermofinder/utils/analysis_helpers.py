"""
Helper functions for analysis output (CSV rows, Excel tabs).
Column maps are built dynamically from the configured temperatures.
"""
from typing import List



_DEFAULT_TEMPS = [25, 37, 42]


def calculate_composition(sequence: str) -> dict:
    """Calculate AU%, GC%, GU% for an RNA sequence."""
    sequence = sequence.upper().replace("T", "U")
    length = len(sequence)

    if length == 0:
        return {"AU%": 0.0, "GC%": 0.0, "GU%": 0.0}

    a_count = sequence.count('A')
    u_count = sequence.count('U')
    g_count = sequence.count('G')
    c_count = sequence.count('C')

    au_percent = ((a_count + u_count) / length) * 100
    gc_percent = ((g_count + c_count) / length) * 100
    gu_percent = ((g_count + u_count) / length) * 100

    return {
        "AU%": round(au_percent, 2),
        "GC%": round(gc_percent, 2),
        "GU%": round(gu_percent, 2)
    }



def _build_column_map(temps: List[int]) -> list:
    """Ordered (setting_key, data_key) pairs for all output columns."""
    base = temps[0]
    t_first = temps[0]
    cmap = []

    cmap += [("name", "name"),
             ("original_sequence", "original_sequence"),
             ("original_structure", "original_structure")]

    for t in temps:
        if t != base:
            cmap.append((f"full_structure_{t}", f"full_structure_{t}"))

    for t in temps:
        cmap.append((f"original_mfe_{t}", f"original_mfe_{t}"))

    cmap += [("original_au_percent", "original_au_percent"),
             ("original_gc_percent", "original_gc_percent"),
             ("original_gu_percent", "original_gu_percent")]

    for t in temps:
        cmap.append((f"original_mfe_{t}_in_range", f"original_mfe_{t}_in_range"))
    cmap += [("original_au_in_range", "original_au_in_range"),
             ("original_gc_in_range", "original_gc_in_range"),
             ("original_gu_in_range", "original_gu_in_range")]

    for t in temps:
        cmap += [(f"full_rbs_{t}_seq",    f"full_rbs_{t}_seq"),
                 (f"full_rbs_{t}_struct",  f"full_rbs_{t}_struct"),
                 (f"full_rbs_{t}_paired",  f"full_rbs_{t}_paired")]
    if len(temps) >= 2:
        cmap.append((f"rbs_seq_diff_{temps[-1]}_{t_first}",
                     f"rbs_seq_diff_{temps[-1]}_{t_first}"))
    if len(temps) >= 3:
        cmap.append((f"rbs_seq_diff_{temps[-2]}_{t_first}",
                     f"rbs_seq_diff_{temps[-2]}_{t_first}"))

    cmap += [("hairpin_detection_method", "hairpin_detection_method"),
             ("rbs_detection_params", "rbs_detection_params"),
             ("hairpin_sequence", "hairpin_sequence"),
             ("hairpin_structure", "hairpin_structure"),
             ("hairpin_au_percent", "hairpin_au_percent"),
             ("hairpin_gc_percent", "hairpin_gc_percent"),
             ("hairpin_gu_percent", "hairpin_gu_percent")]

    for t in temps:
        cmap.append((f"mfe_{t}c_hairpin", f"mfe_{t}c_hairpin"))
    for t in temps:
        cmap.append((f"mfe_{t}_in_range_hairpin", f"mfe_{t}_in_range_hairpin"))

    cmap += [("au_in_range_hairpin", "au_in_range_hairpin"),
             ("gc_in_range_hairpin", "gc_in_range_hairpin"),
             ("gu_in_range_hairpin", "gu_in_range_hairpin")]

    cmap += [("rbs_sequence", "rbs_sequence"),
             ("rbs_structure", "rbs_structure"),
             ("rbs_paired_percent", "rbs_paired_percent")]

    for t in temps:
        cmap.append((f"pf_full_ensemble_{t}", f"pf_full_ensemble_{t}"))
    for t in temps:
        cmap.append((f"pf_full_mean_paired_{t}", f"pf_full_mean_paired_{t}"))

    for t in temps:
        cmap.append((f"pf_hp_ensemble_{t}", f"pf_hp_ensemble_{t}"))
    for t in temps:
        cmap.append((f"pf_hp_mean_paired_{t}", f"pf_hp_mean_paired_{t}"))

    for t in temps:
        cmap.append((f"pf_rbs_access_{t}", f"pf_rbs_access_{t}"))
    if len(temps) >= 2:
        cmap.append((f"pf_rbs_diff_{temps[-1]}_{t_first}",
                     f"pf_rbs_diff_{temps[-1]}_{t_first}"))
    if len(temps) >= 3:
        cmap.append((f"pf_rbs_diff_{temps[-2]}_{t_first}",
                     f"pf_rbs_diff_{temps[-2]}_{t_first}"))

    for t in temps:
        cmap.append((f"pf_hp_ensemble_{t}_in_range", f"pf_hp_ensemble_{t}_in_range"))

    cmap.append(("rbs_paired_in_range", "rbs_paired_in_range"))

    # Motif finder
    cmap += [("motif_pattern", "motif_pattern"),
             ("motif_count", "motif_count"),
             ("motif_match_seq", "motif_match_seq"),
             ("motif_match_pos", "motif_match_pos")]
    for t in temps:
        cmap += [(f"motif_paired_pct_{t}", f"motif_paired_pct_{t}"),
                 (f"motif_struct_{t}", f"motif_struct_{t}"),
                 (f"motif_pf_access_{t}", f"motif_pf_access_{t}")]
    if len(temps) >= 2:
        cmap += [(f"motif_paired_diff_{temps[-1]}_{t_first}",
                  f"motif_paired_diff_{temps[-1]}_{t_first}"),
                 (f"motif_pf_diff_{temps[-1]}_{t_first}",
                  f"motif_pf_diff_{temps[-1]}_{t_first}")]
    if len(temps) >= 3:
        cmap += [(f"motif_paired_diff_{temps[-2]}_{t_first}",
                  f"motif_paired_diff_{temps[-2]}_{t_first}"),
                 (f"motif_pf_diff_{temps[-2]}_{t_first}",
                  f"motif_pf_diff_{temps[-2]}_{t_first}")]

    cmap += [("hp_quality_score", "hp_quality_score"),
             ("hp_quality_score_weighted", "hp_quality_score_weighted"),
             ("hp_quality_score_class", "hp_quality_score_class"),
             ("hp_quality_score_breakdown", "hp_quality_score_breakdown")]

    cmap += [("fl_quality_score", "fl_quality_score"),
             ("fl_quality_score_weighted", "fl_quality_score_weighted"),
             ("fl_quality_score_class", "fl_quality_score_class"),
             ("fl_quality_score_breakdown", "fl_quality_score_breakdown")]

    return cmap


def build_csv_row(result_data: dict, settings_manager) -> list:
    """Build a CSV row from result_data, filtered by enabled columns."""
    column_settings = settings_manager.settings["csv_output_columns"]
    row = []

    temps = _DEFAULT_TEMPS
    if hasattr(settings_manager, 'get_temperatures'):
        temps = settings_manager.get_temperatures()

    column_map = _build_column_map(temps)

    for setting_key, data_key in column_map:
        if column_settings.get(setting_key, False):
            row.append(result_data.get(data_key, ""))

    return row


def format_mfe_value(mfe: float, min_val: float = -15, max_val: float = -5) -> str:
    """Format MFE value or 'Not in Range' if outside thresholds."""
    if min_val <= mfe <= max_val:
        return f"{mfe:.2f}"
    else:
        return "Not in Range"



def _build_full_seq_excel_columns(temps: List[int]) -> list:
    """Column defs for the Full Sequence Excel tab."""
    base = temps[0]
    t_first = temps[0]
    cols = [
        ("name",               "Name"),
        ("original_sequence",  "Sequence"),
        ("original_structure", f"Structure_{base}C"),
    ]
    for t in temps:
        if t != base:
            cols.append((f"full_structure_{t}", f"Structure_{t}C"))
    for t in temps:
        cols.append((f"original_mfe_{t}", f"MFE_{t}C"))
    cols += [
        ("original_au_percent", "AU%"),
        ("original_gc_percent", "GC%"),
        ("original_gu_percent", "GU%"),
    ]
    for t in temps:
        cols.append((f"original_mfe_{t}_in_range", f"MFE_{t}C_InRange"))
    cols += [
        ("original_au_in_range", "AU%_InRange"),
        ("original_gc_in_range", "GC%_InRange"),
        ("original_gu_in_range", "GU%_InRange"),
    ]
    for t in temps:
        cols += [(f"full_rbs_{t}_seq",    f"RBS_{t}C_Seq"),
                 (f"full_rbs_{t}_struct",  f"RBS_{t}C_Struct"),
                 (f"full_rbs_{t}_paired",  f"RBS_{t}C_Paired%")]
    if len(temps) >= 2:
        cols.append((f"rbs_seq_diff_{temps[-1]}_{t_first}", f"RBS_Diff_{temps[-1]}-{t_first}"))
    if len(temps) >= 3:
        cols.append((f"rbs_seq_diff_{temps[-2]}_{t_first}", f"RBS_Diff_{temps[-2]}-{t_first}"))
    for t in temps:
        cols.append((f"pf_full_ensemble_{t}", f"PF_Ensemble_{t}C"))
    for t in temps:
        cols.append((f"pf_full_mean_paired_{t}", f"PF_MeanPaired_{t}C"))
    for t in temps:
        cols.append((f"pf_rbs_access_{t}", f"PF_RBS_Access_{t}C"))
    if len(temps) >= 2:
        cols.append((f"pf_rbs_diff_{temps[-1]}_{t_first}", f"PF_RBS_Diff_{temps[-1]}-{t_first}"))
    if len(temps) >= 3:
        cols.append((f"pf_rbs_diff_{temps[-2]}_{t_first}", f"PF_RBS_Diff_{temps[-2]}-{t_first}"))
    cols += [("motif_pattern",    "Motif_Pattern"),
             ("motif_count",      "Motif_Count"),
             ("motif_match_seq",  "Motif_Match_Seq"),
             ("motif_match_pos",  "Motif_Match_Pos")]
    for t in temps:
        cols += [(f"motif_paired_pct_{t}", f"Motif_Paired%_{t}C"),
                 (f"motif_struct_{t}",     f"Motif_Struct_{t}C"),
                 (f"motif_pf_access_{t}",  f"Motif_PF_Access_{t}C")]
    if len(temps) >= 2:
        cols += [(f"motif_paired_diff_{temps[-1]}_{t_first}", f"Motif_Paired_Diff_{temps[-1]}-{t_first}"),
                 (f"motif_pf_diff_{temps[-1]}_{t_first}",    f"Motif_PF_Diff_{temps[-1]}-{t_first}")]
    if len(temps) >= 3:
        cols += [(f"motif_paired_diff_{temps[-2]}_{t_first}", f"Motif_Paired_Diff_{temps[-2]}-{t_first}"),
                 (f"motif_pf_diff_{temps[-2]}_{t_first}",    f"Motif_PF_Diff_{temps[-2]}-{t_first}")]
    cols += [
        ("fl_quality_score",           "FL_Quality_Score"),
        ("fl_quality_score_weighted",  "FL_Quality_Score_Weighted"),
        ("fl_quality_score_class",     "FL_Quality_Score_Class"),
        ("fl_quality_score_breakdown", "FL_Quality_Score_Breakdown"),
    ]
    return cols


def _build_hairpin_excel_columns(temps: List[int]) -> list:
    """Column defs for the Hairpin Analysis Excel tab."""
    t_first = temps[0]
    cols = [
        ("name",                     "Name"),
        ("hairpin_detection_method", "Detection_Method"),
        ("hairpin_sequence",         "Hairpin_Sequence"),
        ("hairpin_structure",        "Hairpin_Structure"),
        ("hairpin_au_percent",       "AU%"),
        ("hairpin_gc_percent",       "GC%"),
        ("hairpin_gu_percent",       "GU%"),
    ]
    for t in temps:
        cols.append((f"mfe_{t}c_hairpin", f"MFE_{t}C"))
    for t in temps:
        cols.append((f"mfe_{t}_in_range_hairpin", f"MFE_{t}C_InRange"))
    cols += [
        ("au_in_range_hairpin", "AU%_InRange"),
        ("gc_in_range_hairpin", "GC%_InRange"),
        ("gu_in_range_hairpin", "GU%_InRange"),
        ("rbs_sequence",        "RBS_Sequence"),
        ("rbs_structure",       "RBS_Structure"),
        ("rbs_paired_percent",  "RBS_Paired%"),
    ]
    for t in temps:
        cols.append((f"pf_hp_ensemble_{t}", f"PF_Ensemble_{t}C"))
    for t in temps:
        cols.append((f"pf_hp_mean_paired_{t}", f"PF_MeanPaired_{t}C"))
    for t in temps:
        cols.append((f"pf_hp_ensemble_{t}_in_range", f"PF_Ensemble_{t}C_InRange"))
    cols.append(("rbs_paired_in_range", "RBS_Paired%_InRange"))
    cols += [("motif_pattern",    "Motif_Pattern"),
             ("motif_count",      "Motif_Count"),
             ("motif_match_seq",  "Motif_Match_Seq"),
             ("motif_match_pos",  "Motif_Match_Pos")]
    for t in temps:
        cols += [(f"motif_paired_pct_{t}", f"Motif_Paired%_{t}C"),
                 (f"motif_struct_{t}",     f"Motif_Struct_{t}C")]
    if len(temps) >= 2:
        cols.append((f"motif_paired_diff_{temps[-1]}_{t_first}", f"Motif_Paired_Diff_{temps[-1]}-{t_first}"))
    if len(temps) >= 3:
        cols.append((f"motif_paired_diff_{temps[-2]}_{t_first}", f"Motif_Paired_Diff_{temps[-2]}-{t_first}"))
    cols += [
        ("hp_quality_score",           "HP_Quality_Score"),
        ("hp_quality_score_weighted",  "HP_Quality_Score_Weighted"),
        ("hp_quality_score_class",     "HP_Quality_Score_Class"),
        ("hp_quality_score_breakdown", "HP_Quality_Score_Breakdown"),
    ]
    return cols


def _build_motif_excel_columns(temps: List[int]) -> list:
    """Column defs for the Motif Matches Excel tab (one row per hit)."""
    t_first = temps[0]
    cols = [
        ("name",         "Name"),
        ("motif_pattern", "Motif_Pattern"),
        ("match_num",    "Match_#"),
        ("matched_seq",  "Match_Seq"),
        ("match_pos",    "Match_Pos"),
    ]
    for t in temps:
        cols.append((f"mfe_paired_pct_{t}", f"Paired%_{t}C"))
    for t in temps:
        cols.append((f"mfe_struct_{t}", f"Struct_{t}C"))
    for t in temps:
        cols.append((f"pf_access_pct_{t}", f"PF_Access_{t}C"))
    if len(temps) >= 2:
        cols.append((f"mfe_paired_diff_{temps[-1]}_{t_first}",
                     f"Paired_Diff_{temps[-1]}-{t_first}"))
        cols.append((f"pf_access_diff_{temps[-1]}_{t_first}",
                     f"PF_Diff_{temps[-1]}-{t_first}"))
    if len(temps) >= 3:
        cols.append((f"mfe_paired_diff_{temps[-2]}_{t_first}",
                     f"Paired_Diff_{temps[-2]}-{t_first}"))
        cols.append((f"pf_access_diff_{temps[-2]}_{t_first}",
                     f"PF_Diff_{temps[-2]}-{t_first}"))
    return cols


def _expand_motif_hits(results: list, temps: List[int]) -> list:
    """Flatten _motif_hits from each result into one row per hit."""
    rows = []
    for result_data in results:
        motif_hits = result_data.get("_motif_hits")
        if not motif_hits:
            continue
        name = result_data.get("name", "")
        pattern = result_data.get("motif_pattern", "")
        for i, hit in enumerate(motif_hits, start=1):
            row = {
                "name": name,
                "motif_pattern": pattern,
                "match_num": i,
                "matched_seq": hit.get("matched_seq", ""),
                "match_pos": f"{hit.get('start', '')}-{hit.get('end', '')}",
            }
            for t in temps:
                pct = hit.get(f"mfe_paired_pct_{t}")
                row[f"mfe_paired_pct_{t}"] = f"{pct:.2f}" if pct is not None else "N/A"
                st = hit.get(f"mfe_struct_{t}")
                row[f"mfe_struct_{t}"] = st if st else "N/A"
                acc = hit.get(f"pf_access_pct_{t}")
                row[f"pf_access_pct_{t}"] = f"{acc:.2f}" if acc is not None else "N/A"
            t_first = temps[0]
            if len(temps) >= 2:
                d = hit.get(f"mfe_paired_diff_{temps[-1]}_{t_first}")
                row[f"mfe_paired_diff_{temps[-1]}_{t_first}"] = (
                    f"{d:+.2f}" if d is not None else "N/A")
                d2 = hit.get(f"pf_access_diff_{temps[-1]}_{t_first}")
                row[f"pf_access_diff_{temps[-1]}_{t_first}"] = (
                    f"{d2:+.2f}" if d2 is not None else "N/A")
            if len(temps) >= 3:
                d = hit.get(f"mfe_paired_diff_{temps[-2]}_{t_first}")
                row[f"mfe_paired_diff_{temps[-2]}_{t_first}"] = (
                    f"{d:+.2f}" if d is not None else "N/A")
                d2 = hit.get(f"pf_access_diff_{temps[-2]}_{t_first}")
                row[f"pf_access_diff_{temps[-2]}_{t_first}"] = (
                    f"{d2:+.2f}" if d2 is not None else "N/A")
            rows.append(row)
    return rows


def write_excel_with_tabs(results: list, output_path, settings_manager=None,
                          temps: List[int] = None):
    """Write results to Excel with Full Sequence, Hairpin Analysis,
    and (if applicable) Motif Matches tabs."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if temps is None:
        if settings_manager and hasattr(settings_manager, 'get_temperatures'):
            temps = settings_manager.get_temperatures()
        else:
            temps = list(_DEFAULT_TEMPS)

    wb = Workbook()

    full_seq_columns = _build_full_seq_excel_columns(temps)
    hairpin_columns = _build_hairpin_excel_columns(temps)

    if settings_manager:
        col_settings = settings_manager.settings.get("csv_output_columns", {})
        full_seq_columns = [
            (k, h) for k, h in full_seq_columns if col_settings.get(k, False)
        ]
        hairpin_columns = [
            (k, h) for k, h in hairpin_columns if col_settings.get(k, False)
        ]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def _write_sheet(ws, columns, data_rows):
        for col_idx, (_, display_name) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, result_data in enumerate(data_rows, start=2):
            for col_idx, (data_key, _) in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=result_data.get(data_key, ""))

        # Auto-fit widths
        for col_idx, (_, display_name) in enumerate(columns, start=1):
            max_len = len(display_name)
            for row_idx in range(2, min(len(data_rows) + 2, 52)):  # sample first 50 rows
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    ws_full = wb.active
    ws_full.title = "Full Sequence"
    if full_seq_columns:
        _write_sheet(ws_full, full_seq_columns, results)

    ws_hairpin = wb.create_sheet(title="Hairpin Analysis")
    if hairpin_columns:
        _write_sheet(ws_hairpin, hairpin_columns, results)

    motif_rows = _expand_motif_hits(results, temps)
    if motif_rows:
        motif_columns = _build_motif_excel_columns(temps)
        ws_motif = wb.create_sheet(title="Motif Matches")
        _write_sheet(ws_motif, motif_columns, motif_rows)

    wb.save(output_path)
